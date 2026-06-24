from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from datetime import datetime
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import time

# =========================
# 1. Configurações
# =========================

URL = "https://www.nescafe-dolcegusto.com.br/sabores"
OUTPUT_DIR = Path(__file__).parent
TODAY = datetime.now().strftime("%Y-%m-%d")
OUTPUT_FILE = OUTPUT_DIR / f"{TODAY}_dolce_gusto_capsulas_precos.xlsx"

# =========================
# 2. Configurar navegador
# =========================

options = Options()
options.add_argument("--headless=new")   # headless moderno — renderiza JS igual ao browser normal
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-gpu")
options.add_argument("--window-size=1920,1080")
options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("useAutomationExtension", False)

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)
driver.set_page_load_timeout(60)

# =========================
# 3. Abrir site e aguardar produtos
# =========================

print("🔍 Acessando site da Dolce Gusto...")
driver.get(URL)

# Aguarda os produtos aparecerem (até 40s)
try:
    WebDriverWait(driver, 40).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".spc-product__name"))
    )
    print("✅ Produtos detectados na página.")
except Exception:
    print("⚠️  Timeout — salvando HTML para diagnóstico e tentando continuar...")
    with open("debug_page.html", "w", encoding="utf-8") as f:
        f.write(driver.page_source)

# =========================
# 4. Aceitar cookies, se aparecer
# =========================

try:
    botoes = driver.find_elements(By.TAG_NAME, "button")
    for botao in botoes:
        texto = botao.text.strip().lower()
        if any(p in texto for p in ["aceitar", "accept", "concordo", "permitir"]):
            botao.click()
            time.sleep(1)
            break
except Exception:
    pass

# =========================
# 5. Rolar página para carregar lazy load
# =========================

last_height = driver.execute_script("return document.body.scrollHeight")

for _ in range(20):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)
    new_height = driver.execute_script("return document.body.scrollHeight")
    if new_height == last_height:
        break
    last_height = new_height

# =========================
# 6. Capturar HTML e parsear
# =========================

soup = BeautifulSoup(driver.page_source, "html.parser")
driver.quit()

# =========================
# 7. Extrair produtos
# =========================

cards = soup.select(".spc-product")
print(f"Produtos encontrados: {len(cards)}")

results = []

for card in cards:
    # Nome — <h2 class="spc-product__name" data-automation="product-title">
    name_tag = card.select_one("[data-automation='product-title']")
    product_name = name_tag.get_text(strip=True) if name_tag else ""

    # Preço — <span class="spc-product__price--regular" data-automation="product-price">
    price_tag = card.select_one("[data-automation='product-price']")
    current_price_text = price_tag.get_text(strip=True) if price_tag else ""

    # Limpa o preço: "R$2,79" → 2.79
    price_clean = re.sub(r"[^\d,]", "", current_price_text).replace(",", ".")
    try:
        current_price_numeric = float(price_clean)
    except ValueError:
        current_price_numeric = None

    # Quantidade de cápsulas — <div class="spc-product__pods-qty">
    pods_tag = card.select_one(".spc-product__pods-qty")
    pods_text = pods_tag.get_text(strip=True) if pods_tag else ""
    pods_match = re.search(r"(\d+)", pods_text)
    pods_qty = int(pods_match.group(1)) if pods_match else None

    # SKU — data-sku no card pai .spc-product
    sku = card.get("data-sku", "")

    # Categoria — pelo container pai .spc-listing__container[name]
    cat_container = card.find_parent(attrs={"name": True})
    category = cat_container.get("name", "").strip() if cat_container else ""

    # Custo por cápsula
    custo_por_capsula = round(current_price_numeric / pods_qty, 4) if current_price_numeric and pods_qty else None

    if product_name or current_price_text:
        results.append({
            "data_atualizacao": datetime.now().strftime("%d/%m/%Y"),
            "categoria": category,
            "produto": product_name,
            "preco": current_price_text,
            "preco_numerico": current_price_numeric,
            "capsulas": pods_qty,
            "custo_por_capsula": custo_por_capsula,
            "sku": sku,
        })

# =========================
# 8. Montar DataFrame
# =========================

df = pd.DataFrame(results)

if df.empty:
    print("❌ Nenhum produto extraído. Verifique o arquivo debug_page.html no artifact.")
    exit(1)

df["preco_numerico"] = pd.to_numeric(df["preco_numerico"], errors="coerce")
df["custo_por_capsula"] = pd.to_numeric(df["custo_por_capsula"], errors="coerce")

print(f"✅ {len(df)} produtos extraídos.")

# =========================
# 9. Gerar Excel formatado
# =========================

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cápsulas"

hd_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
hd_fill   = PatternFill("solid", start_color="B00000")
data_font = Font(name="Arial", size=10)
center    = Alignment(horizontal="center", vertical="center")
left_al   = Alignment(horizontal="left", vertical="center")
brl_fmt   = 'R$#,##0.00'
thin      = Side(style="thin", color="DDDDDD")
border    = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.merge_cells("A1:H1")
ws["A1"] = "Nescafé Dolce Gusto — Preços das Cápsulas"
ws["A1"].font = Font(name="Arial", bold=True, size=13, color="B00000")
ws["A1"].alignment = center

ws.merge_cells("A2:H2")
ws["A2"] = f"Atualizado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
ws["A2"].font = Font(name="Arial", size=9, color="888888")
ws["A2"].alignment = center

ws.row_dimensions[1].height = 24
ws.row_dimensions[2].height = 16

headers = [
    ("#", 5), ("Data", 13), ("Categoria", 18), ("Produto", 44),
    ("Preço", 14), ("Cápsulas", 12), ("Custo/Cápsula", 15), ("SKU", 14),
]

for col, (h, w) in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = hd_font
    cell.fill = hd_fill
    cell.alignment = center
    cell.border = border
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[3].height = 20

for i, row_data in enumerate(df.itertuples(index=False), 1):
    r = i + 3
    alt_fill = PatternFill("solid", start_color="FFF5F5" if i % 2 == 0 else "FFFFFF")

    values = [
        i, row_data.data_atualizacao, row_data.categoria, row_data.produto,
        row_data.preco_numerico, row_data.capsulas, row_data.custo_por_capsula, row_data.sku,
    ]
    aligns = [center, center, center, left_al, center, center, center, center]
    fmts   = [None, None, None, None, brl_fmt, None, brl_fmt, None]

    for col, (val, aln, fmt) in enumerate(zip(values, aligns, fmts), 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = data_font
        cell.fill = alt_fill
        cell.alignment = aln
        cell.border = border
        if fmt and val is not None:
            cell.number_format = fmt

    ws.row_dimensions[r].height = 17

ws.freeze_panes = "A4"
wb.save(OUTPUT_FILE)
print(f"📊 Planilha salva em: {OUTPUT_FILE}")